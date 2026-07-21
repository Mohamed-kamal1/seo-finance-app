"""Import one Excel workbook into SEO House Finance.

Each worksheet represents an app page/table. Sheet names are case-insensitive and
may use spaces, hyphens, or underscores. For example: ``Clients``, ``Client
Balances``, ``Invoices``, ``Transactions``, ``Treasuries``, ``Guest Posts``,
``Guest Post Ledger``, ``Content Billing``, and ``Content Details``.

The first non-empty row in each sheet is treated as its header. Headers may use
database column names (``actual_date``) or the friendly forms used by the app
(``Actual Date``, ``Client Name``, ``Treasury``, and ``Site``). Unknown sheets
and columns are reported and skipped, which makes it safe to keep notes tabs in
the same workbook.

Usage:
    pip install -r requirements.txt
    python import_data.py /path/to/finance.xlsx --dry-run
    python import_data.py /path/to/finance.xlsx

Set SUPABASE_URL (or NEXT_PUBLIC_SUPABASE_URL) and SUPABASE_SERVICE_ROLE_KEY
before a non-dry-run import.
"""

import argparse
import os
import re
import sys
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from dotenv import load_dotenv
from supabase import create_client


load_dotenv(".env.local")

BATCH_SIZE = 500
NUM_RE = re.compile(r"[-+]?\d[\d,.]*")

# Worksheet names accepted by the importer. Add aliases here when a workbook
# uses a new page label; the database table names remain deliberately explicit.
SHEET_TABLES = {
    "currencies": "currencies",
    "clients": "clients",
    "client balances": "client_balances",
    "balances": "client_balances",
    "invoices": "invoices",
    "manual invoices": "invoices",
    "chart of accounts": "chart_of_accounts",
    "classifications": "classifications",
    "treasuries": "treasury_accounts",
    "treasury accounts": "treasury_accounts",
    "transactions": "transactions",
    "ledger": "transactions",
    "guest posts": "guest_post_sites",
    "guest post sites": "guest_post_sites",
    "guest post ledger": "guest_post_ledger",
    "content billing": "content_billing",
    "content details": "content_details",
}

TABLE_COLUMNS = {
    "currencies": {"code", "name", "symbol", "rate_to_base", "is_base"},
    "clients": {"name", "website", "country", "payment_duration", "currency_code", "seo_fee", "guest_fee", "hosting_fee", "content_fee", "annual_increase", "increase_applies_date", "contract_date", "billing_day", "service_type", "notes", "status"},
    "client_balances": {"client_id", "as_of_date", "seo", "guest", "hosting_domain", "content", "past_due", "discount", "total_amount", "collections", "current_due", "currency_code", "notes"},
    "invoices": {"internal_id", "client_id", "invoice_date", "service", "seo", "guest", "hosting_domain", "content", "past_due", "discount", "total_amount", "collections", "current_due", "currency_code", "collection_status", "payment_date", "notes"},
    "chart_of_accounts": {"category", "group_type"},
    "classifications": {"name"},
    "treasury_accounts": {"name", "currency_code", "opening_balance", "opening_date", "notes"},
    "transactions": {"actual_date", "cf_date", "description", "notes", "debit", "credit", "classification_is", "classification_cf", "treasury_account_id", "statement", "source"},
    "guest_post_sites": {"name", "client_id", "website_url"},
    "guest_post_ledger": {"site_id", "month", "beg_balance", "credit", "content", "transfer", "current_balance"},
    "content_billing": {"client_id", "client_name_raw", "details", "required_amount", "paid_amount", "balance", "currency_code", "period", "notes"},
    "content_details": {"words", "price", "currency_code"},
}

HEADER_ALIASES = {
    "client": "client_name", "client name": "client_name",
    "treasury": "treasury_name", "treasury name": "treasury_name",
    "site": "site_name", "site name": "site_name",
    "website url": "website_url", "hosting domain": "hosting_domain",
    "actual date": "actual_date", "cf date": "cf_date",
    "invoice date": "invoice_date", "payment date": "payment_date",
    "as of date": "as_of_date", "opening date": "opening_date",
    "required amount": "required_amount", "paid amount": "paid_amount",
    "current due": "current_due", "total amount": "total_amount",
    "collection status": "collection_status", "currency": "currency_code",
    "currency code": "currency_code", "group type": "group_type",
    "beg balance": "beg_balance", "current balance": "current_balance",
    "classifications is": "classification_is", "classifications cf": "classification_cf",
}

NUMERIC_COLUMNS = {
    "rate_to_base", "seo_fee", "guest_fee", "hosting_fee", "content_fee",
    "annual_increase", "seo", "guest", "hosting_domain", "content", "past_due",
    "discount", "total_amount", "collections", "current_due", "opening_balance",
    "debit", "credit", "beg_balance", "transfer", "current_balance",
    "required_amount", "paid_amount", "balance", "price",
}
DATE_COLUMNS = {"increase_applies_date", "contract_date", "as_of_date", "invoice_date", "payment_date", "opening_date", "actual_date", "cf_date", "month", "period"}


def normalise(value):
    """Make a sheet/header label comparable without changing user data."""
    return re.sub(r"\s+", " ", re.sub(r"[_-]+", " ", str(value).strip().lower()))


def column_name(header):
    label = normalise(header)
    return HEADER_ALIASES.get(label, label.replace(" ", "_"))


def parse_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    match = NUM_RE.search(str(value))
    return float(match.group().replace(",", "")) if match else None


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date().isoformat()
            except ValueError:
                pass
    return None


def parse_bool(value):
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "yes", "1", "y"}


def header_row_index(ws):
    for row_index, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 25), values_only=True), 1):
        if any(value not in (None, "") for value in row):
            return row_index
    return None


def read_sheet(ws, table):
    """Return valid records plus the raw relationship labels for a worksheet."""
    header_index = header_row_index(ws)
    if header_index is None:
        return [], set()
    headers = [column_name(cell.value) if cell.value not in (None, "") else None for cell in ws[header_index]]
    unknown = {header for header in headers if header and header not in TABLE_COLUMNS[table] and header not in {"client_name", "treasury_name", "site_name"}}
    records = []
    for values in ws.iter_rows(min_row=header_index + 1, values_only=True):
        raw = {header: values[index] for index, header in enumerate(headers) if header and index < len(values) and values[index] not in (None, "")}
        if not raw:
            continue
        row = {}
        for key, value in raw.items():
            if key in NUMERIC_COLUMNS:
                value = parse_number(value)
            elif key in DATE_COLUMNS:
                value = parse_date(value)
            elif key == "is_base":
                value = parse_bool(value)
            elif isinstance(value, str):
                value = value.strip()
            if value is not None:
                row[key] = value
        if row:
            records.append(row)
    return records, unknown


def lookup_by_name(supabase, table, label_column="name"):
    rows = supabase.table(table).select(f"id,{label_column}").execute().data or []
    return {str(row[label_column]).strip(): row["id"] for row in rows if row.get(label_column) is not None}


def insert_batches(supabase, table, rows, conflict=None):
    for start in range(0, len(rows), BATCH_SIZE):
        query = supabase.table(table)
        batch = rows[start:start + BATCH_SIZE]
        if conflict:
            query.upsert(batch, on_conflict=conflict).execute()
        else:
            query.insert(batch).execute()


def main():
    parser = argparse.ArgumentParser(description="Import page-named worksheets from one .xlsx file.")
    parser.add_argument("workbook", help="Path to the Excel workbook (.xlsx)")
    parser.add_argument("--dry-run", action="store_true", help="Validate and show import counts without writing data")
    args = parser.parse_args()

    if not args.workbook.lower().endswith(".xlsx"):
        parser.error("workbook must be an .xlsx file")
    if not os.path.isfile(args.workbook):
        parser.error(f"workbook not found: {args.workbook}")

    workbook = openpyxl.load_workbook(args.workbook, data_only=True)
    imported = {}
    skipped = []
    for ws in workbook.worksheets:
        table = SHEET_TABLES.get(normalise(ws.title))
        if not table:
            skipped.append(ws.title)
            continue
        rows, unknown = read_sheet(ws, table)
        imported.setdefault(table, []).extend(rows)
        print(f"{ws.title}: parsed {len(rows)} row(s) for {table}." + (f" Ignored columns: {', '.join(sorted(unknown))}." if unknown else ""))

    if skipped:
        print(f"Skipped unrecognised sheet(s): {', '.join(skipped)}.")
    if not imported:
        raise ValueError("No recognised, non-empty worksheets found. See the script docstring for supported names.")
    if args.dry_run:
        print("Dry run complete: no data was written.")
        return

    url = os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        print("Set SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY before importing.")
        sys.exit(1)
    supabase = create_client(url, key)

    # Insert independent lookup tables first. Their natural keys make repeated
    # imports safe for these entities.
    for table, conflict in (("currencies", "code"), ("chart_of_accounts", "category,group_type"), ("classifications", "name"), ("treasury_accounts", "name")):
        rows = imported.get(table, [])
        if rows:
            insert_batches(supabase, table, rows, conflict)
            print(f"Imported {len(rows)} {table} row(s).")

    client_rows = imported.get("clients", [])
    if client_rows:
        existing_clients = lookup_by_name(supabase, "clients")
        new_clients = [row for row in client_rows if row.get("name") and row["name"] not in existing_clients]
        insert_batches(supabase, "clients", new_clients)
        print(f"Imported {len(new_clients)} new client row(s); skipped {len(client_rows) - len(new_clients)} existing client(s).")

    client_ids = lookup_by_name(supabase, "clients")
    treasury_ids = lookup_by_name(supabase, "treasury_accounts")

    site_rows = imported.get("guest_post_sites", [])
    for row in site_rows:
        client_name = row.pop("client_name", None)
        if client_name:
            row["client_id"] = client_ids.get(str(client_name))
    if site_rows:
        insert_batches(supabase, "guest_post_sites", site_rows, "name")
        print(f"Imported {len(site_rows)} guest_post_sites row(s).")
    site_ids = lookup_by_name(supabase, "guest_post_sites")

    for table in ("client_balances", "invoices", "transactions", "guest_post_ledger", "content_billing", "content_details"):
        prepared = []
        for source_row in imported.get(table, []):
            row = dict(source_row)
            client_name = row.pop("client_name", None)
            treasury_name = row.pop("treasury_name", None)
            site_name = row.pop("site_name", None)
            if client_name:
                row["client_id"] = client_ids.get(str(client_name))
                if table == "content_billing":
                    row.setdefault("client_name_raw", str(client_name))
            if treasury_name:
                row["treasury_account_id"] = treasury_ids.get(str(treasury_name))
            if site_name:
                row["site_id"] = site_ids.get(str(site_name))
            missing = {"client_id": client_name, "treasury_account_id": treasury_name, "site_id": site_name}
            unresolved = [label for column, label in missing.items() if label and not row.get(column)]
            if unresolved:
                print(f"Skipping {table} row with unresolved reference(s): {', '.join(map(str, unresolved))}.")
                continue
            prepared.append(row)
        if prepared:
            conflict = "site_id,month" if table == "guest_post_ledger" else None
            insert_batches(supabase, table, prepared, conflict)
            print(f"Imported {len(prepared)} {table} row(s).")

    print("Import complete.")


if __name__ == "__main__":
    main()
